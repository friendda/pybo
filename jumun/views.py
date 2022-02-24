from django.shortcuts import render, get_object_or_404, redirect
from .models import jumun_T
##from .models import Document
from .forms import RegisterForm
from django.utils import timezone
import xlwt
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q
import openpyxl
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
import simplejson as json
from django.contrib import messages




# Create your views here.


##def index(request):
##    return HttpResponse("jumun data")



def index(request):
    #페이지
    page = request.GET.get('page','1')

    #검색어
    kw = request.GET.get('kw', '')
    mkw1 = request.GET.get('mkw1', '').strip()
    mkw2 = request.GET.get('mkw2', '').strip()
    mkw3 = request.GET.get('mkw3', '').strip()
    mkw4 = request.GET.get('mkw4', '').strip()
    mkw5 = request.GET.get('mkw5', '').strip()
    mkc1 = request.GET.get('mkc1', '').strip()
    mkc2 = request.GET.get('mkc2', '').strip()
    mkc3 = request.GET.get('mkc3', '').strip()
    mkc4 = request.GET.get('mkc4', '').strip()
    mkc5 = request.GET.get('mkc5', '').strip()
    
    
    
    #조회
    
    jumun_list = jumun_T.objects.order_by('id')
    
    all_c = jumun_list.count()

    if kw:
        jumun_list = jumun_list.filter(
            Q(jumun_id__icontains=kw) | # 주문id
            Q(jumun_name__icontains=kw) |  # 주문자명
            Q(jumun_con__icontains=kw) |  # 위탁
            Q(brand__icontains=kw)  # 브랜드
        )

    elif mkw1 or mkw2 or mkw3 or mkw4 or mkw5:
        if mkc1 or mkc2 or mkc3 or mkc4 or mkc5:
            jumun_list = jumun_list.filter(
                Q(jumun_name__in=[mkw1,mkw2,mkw3,mkw4,mkw5]),  # 주문자
                Q(jumun_con__in=[mkc1,mkc2,mkc3,mkc4,mkc5])   # 위탁자
            
            )
        else:
            jumun_list = jumun_list.filter(
                Q(jumun_name__in=[mkw1,mkw2,mkw3,mkw4,mkw5]))
                
    kw_c = jumun_list.count()
    #jumun_list2 = 전체 조회내용
    #페이징
    paginator = Paginator(jumun_list, 20) #페이지당 20개
    page_obj = paginator.get_page(page)
    context = {'jumun_list':page_obj, 'jumun_list2':jumun_list, 'page': page, 'kw': kw, 'all_c':all_c, 'kw_c':kw_c,'mkw1':mkw1,'mkw2':mkw2,'mkw3':mkw3,'mkw4':mkw4,'mkw5':mkw5,'mkc1':mkc1,'mkc2':mkc2,'mkc3':mkc3,'mkc4':mkc4,'mkc5':mkc5}
##    context = {'jumun_list': jumun_list}
    return render(request, 'jumun/jumun_list.html', context)

    
    




def jumun_create(request):

 
    if request.method == 'POST':
       form = RegisterForm(request.POST)
       if form.is_valid():
           jumun_T = form.save(commit=False)
           jumun_T.jumun_date = timezone.now()
           jumun_T.save()
           return redirect('jumun:index')
    else:
        form = RegisterForm()
    context = {'form' : form}
    return render(request, 'jumun/jumun_form.html', {'form': form})


def jumun_delete(request):

    id_list2 = request.POST.getlist('selected[]')
    page = request.GET.get('page')    
    for d_id in id_list2:
        d = jumun_T.objects.get(id=d_id)
        d.delete()
    
    return redirect('jumun:index')


## 페이징안된 전체 검색리스트 다운로드
def jumun_excel2(request):
	
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'all_download.xls'


    wb = xlwt.Workbook(encoding='ansi') #encoding은 ansi로 해준다.
    ws = wb.add_sheet('주문장') #시트 추가
    
    row_num = 0
    col_names = ['영문', '주문일','주문ID','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','ID','나온날짜']
    
    #열이름을 첫번째 행에 추가 시켜준다.
    for idx, col_name in enumerate(col_names):
        ws.write(row_num, idx, col_name)


    id_list2 = request.POST.getlist('jumunlist[]')

    
    
    for d_id in id_list2:
        rows = jumun_T.objects.filter(id=d_id).values_list('eng_flag', 'jumun_date','jumun_id','jumun_name','jumun_con','brand','prd_name','prd_color','quantity','wholesale','whole_pr','note','sup_pr','name','phone','address','jumun_id2','date2')
        

        for row in rows:
            row_num +=1
            for col_num, attr in enumerate(row):
                ws.write(row_num, col_num, attr)
    wb.save(response)
    return response
    

##선택 다운로드
def jumun_excel(request):
	
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'select_download.xls'

##    id_list = request.POST.getlist('selected[]')
    wb = xlwt.Workbook(encoding='ansi') #encoding은 ansi로 해준다.
    ws = wb.add_sheet('주문장') #시트 추가
    
    row_num = 0
    col_names = ['영문', '주문일','주문ID','주문자','위탁자','브랜드','상품명','색상','수량','중도매','도매가','비고','공급가','이름','전화번호','주소','ID','나온날짜']
    
    #열이름을 첫번째 행에 추가 시켜준다.
    for idx, col_name in enumerate(col_names):
        ws.write(row_num, idx, col_name)


    id_list2 = request.POST.getlist('selected[]')


    
    
    for d_id in id_list2:
        rows = jumun_T.objects.filter(id=d_id).values_list('eng_flag', 'jumun_date','jumun_id','jumun_name','jumun_con','brand','prd_name','prd_color','quantity','wholesale','whole_pr','note','sup_pr','name','phone','address','jumun_id2','date2')
        

        for row in rows:
            row_num +=1
            for col_num, attr in enumerate(row):
                ws.write(row_num, col_num, attr)
    wb.save(response)
    return response


# 파일업로드 구현부
##def uploadFile(request):
##    if request.method == "POST":
##        # Fetching the form data
##        fileTitle = request.POST["fileTitle"]
##        uploadedFile = request.FILES["uploadedFile"]
##
##        # Saving the information in the database
##        document = Document(
##            title=fileTitle,
##            uploadedFile=uploadedFile
##        )
##        document.save()
##
##    documents = Document.objects.all()
##
####    return render(request, 'jumun/jumun_list.html', context={"files": documents})
##    return redirect('jumun:index')


    
##@csrf_exempt
#엑셀업로드 및 db저장
def excel_upload(request):
    if request.method == 'POST':

        # 파일 저장
        file = request.FILES['file_excel']
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        uploaded_file_url = fs.url(filename)
        # print(uploaded_file_url)

        excel = openpyxl.load_workbook(file, data_only=True)

        work_sheet = excel.worksheets[0]

        #엑셀파일 마지막로우 체크
        cc= work_sheet.max_row
        cc = str(cc)

        rows = work_sheet['A2':'R'+cc]

        for row in rows:
            
            dict = {}
            dict['eng_flag']=row[0].value
            dict['jumun_date']=row[1].value
            dict['jumun_id']=row[2].value
            dict['jumun_name']=row[3].value
            dict['jumun_con']=row[4].value
            dict['brand']=row[5].value
            dict['prd_name']=row[6].value
            dict['prd_color']=row[7].value
            dict['quantity']=row[8].value
            dict['wholesale']=row[9].value
            dict['whole_pr']=row[10].value
            dict['note']=row[11].value
            dict['sup_pr']=row[12].value
            dict['name']=row[13].value
            dict['phone']=row[14].value
            dict['address']=row[15].value
            dict['jumun_id2']=row[16].value
            dict['date2']=row[17].value

            jumun_T(**dict).save()

        context = {'status': True, 'rtnmsg': '엑셀파일이 정상적으로 업로드 됐습니다.'}
        return HttpResponse(json.dumps(context), content_type='application/json')


def all_delete(request):
    jumun_list = jumun_T.objects.all()
    jumun_list.delete()
    messages.info(request, '모든 데이터가 삭제되었습니다.')
    return redirect('jumun:index')

    
